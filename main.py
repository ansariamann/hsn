import os
import openpyxl
from openpyxl.styles import Font, PatternFill
import time
import argparse

from hsn_tree_builder import build_hs_tree
from product_extractor import extract_products_from_url, extract_products_from_text
from hsn_matcher import HSNMatcher

INPUT_EXCEL = 'Book1.xlsx'
OUTPUT_EXCEL = 'hsn_output.xlsx'

def main(argv=None):
    parser = argparse.ArgumentParser(description="HSN Code Lookup Project for Saudi Arabian Companies")
    parser.add_argument("--rebuild-hs-tree", action="store_true", help="Rebuild hs_codes_sa.json even if it exists")
    parser.add_argument("--verbose-hs-tree", action="store_true", help="Verbose HS-tree crawl (prints every leaf code)")
    parser.add_argument(
        "--stop-at-digits",
        type=int,
        default=None,
        help="Stop crawling deeper once HS code reaches N digits (e.g. 6 or 8)",
    )
    args = parser.parse_args(argv)

    print("="*50)
    print("HSN Code Lookup Project for Saudi Arabian Companies")
    print("="*50)
    
    # Step 1: Ensure HS code tree is built and cached
    build_hs_tree(
        force_rebuild=args.rebuild_hs_tree,
        verbose=args.verbose_hs_tree,
        stop_at_digits=args.stop_at_digits,
    )
    
    # Step 2: Initialize HSN Matcher
    matcher = HSNMatcher()
    
    # Step 3: Load input Excel
    if not os.path.exists(INPUT_EXCEL):
        print(f"Error: Could not find input file '{INPUT_EXCEL}'")
        return
        
    print(f"\nLoading data from {INPUT_EXCEL}...")
    wb_in = openpyxl.load_workbook(INPUT_EXCEL)
    ws_in = wb_in.active
    
    # Identify key columns (0-indexed)
    headers = [cell.value for cell in ws_in[1]]
    try:
        col_company = headers.index('Company Name')
        col_products = headers.index('Products name')
        col_url = headers.index('Website URL')
    except ValueError as e:
        print(f"Error finding required columns: {e}")
        return
        
    # Step 4: Prepare output Excel
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "HSN Data"
    
    out_headers = ['Company Name', 'Website URL', 'Product Name', 'HSN Code', 'HSN Description', 'Match Confidence']
    ws_out.append(out_headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, len(out_headers) + 1):
        ws_out.cell(row=1, column=col).font = header_font
        ws_out.cell(row=1, column=col).fill = header_fill
        
    ws_out.column_dimensions['A'].width = 30 # Company
    ws_out.column_dimensions['B'].width = 30 # URL
    ws_out.column_dimensions['C'].width = 40 # Product
    ws_out.column_dimensions['D'].width = 15 # HSN Code
    ws_out.column_dimensions['E'].width = 50 # Description
    ws_out.column_dimensions['F'].width = 15 # Confidence
    
    # Step 5: Process each company
    total_rows = ws_in.max_row
    print(f"\nProcessing {total_rows - 1} companies...")
    
    # Process all rows
    max_process = total_rows
    
    for row_idx in range(2, min(total_rows + 1, max_process + 1)):
        row = [cell.value for cell in ws_in[row_idx]]
        
        company = row[col_company]
        if not company:
            continue
            
        print(f"\n[{row_idx-1}/{min(total_rows-1, max_process-1)}] Company: {company}")
        
        # Extract from URL
        url = row[col_url]
        print(f"  Website: {url}")
        # NOTE: The quality of product extraction from URLs and text is critical.
        # The implementation of `extract_products_from_url` and `extract_products_from_text`
        # in `product_extractor.py` directly impacts how many products are found and matched.
        web_products = extract_products_from_url(url)
        
        # Extract from text fields fallback
        text_products = extract_products_from_text(row[col_products])
        
        # Combine unique products
        all_products = list(set(web_products + text_products))
        
        if not all_products:
            print("  No products found.")
            ws_out.append([company, url, "No products found", "", "", ""])
            continue

        print(f"  Found {len(all_products)} products. Matching HSN codes...")
        
        # Match HSN for EACH product
        for product in all_products:
            if not product or len(str(product).strip()) < 2:
                continue
                
            hs_code, hs_desc, score = matcher.get_match(product)
            
            # Clean description if it redundantly starts with the code.
            # The user noted that descriptions can be redundant. This is a simple cleanup.
            # The root cause of poor descriptions is often a shallow HSN code tree,
            # which results in matching to generic categories.
            if hs_desc and hs_code and str(hs_desc).strip().startswith(str(hs_code)):
                hs_desc = str(hs_desc).replace(str(hs_code), "", 1).strip(" -.")

            print(f"    {product} -> {hs_code} ({score}%)")
            
            ws_out.append([
                company,
                url,
                product,
                hs_code or "Not Found",
                hs_desc or "No Match",
                f"{score}%" if score > 0 else "N/A"
            ])
        
        # Small delay to be nice to servers
        time.sleep(0.5)
        
    # Save output
    output_path = os.path.abspath(OUTPUT_EXCEL)
    wb_out.save(output_path)
    print(f"\nDone! Output saved to: {output_path}")

if __name__ == "__main__":
    main()
